from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import json
import openpyxl

class Test_Tobeto:
    def __init__(self,driver):
        self.driver = driver

    # on ek >> test_    pytestte fonksiyonu çalıştırmak için fonk adı test_ ile başlamalı

    def setup_method(self):   #her test başlangıcında çalışacak fonk.
        self.driver = webdriver.Chrome()
        self.driver.get("https://tobeto.com/giris")
        self.driver.maximize_window()

    def teardown_method(self):  # her testin btiminde çalışacak fonk
        self.driver.quit()

    def getData():
        excel = openpyxl.load_workbook("data/tobetoData.xlsx")
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            email = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((email,password))
        return data
    
    def readInvalidDataFromJson():
        file = open("data/tobetoData.json") 
        data = json.load(file)
        parameter = []

        for user in data['users']:
            email = user["email"]
            password = user["password"]
            parameter.append((email,password))

        return parameter
    
    @pytest.mark.parametrize("email,password",readInvalidDataFromJson())
    def test_yanlis_girdi_basarisiz_giris(self,email,password):
        email_input = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"email")))
        email_input.send_keys(email)
        password_input = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"password")))
        password_input.send_keys(password)
        giris_butonu = self.driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-primary w-100 mt-6']").click()
        mesaj = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class ='toast-body']")))
        assert mesaj.text == "• Geçersiz e-posta veya şifre."

    def test_basarili_giris(self):
        email_input = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.NAME,"email")))
        email_input.send_keys("xabiw41724@talmetry.com")
        password_input = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.NAME,"password")))
        password_input.send_keys("123456")
        giris_butonu = self.driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-primary w-100 mt-6']").click()
        mesaj = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='toast-body']")))
        assert mesaj.text == "• Giriş başarılı."

    def test_bos_alanla_giris(self):
        email_input = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"email")))
        email_input.send_keys("")
        password_input = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"password")))
        password_input.send_keys("")
        giris_butonu = self.driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-primary w-100 mt-6']").click()
        mesaj_1 = self.driver.find_element(By.CSS_SELECTOR,"p[style='text-align: start; color: red;']:nth-child(2)").text
        mesaj_2 = self.driver.find_element(By.CSS_SELECTOR,"p[style='text-align: start; color: red;']:nth-child(4)").text
        assert mesaj_1 == "Doldurulması zorunlu alan*" 
        assert mesaj_2 == "Doldurulması zorunlu alan*"
    
    def test_basarili_sifre_yenileme(self):
        
        sifremi_unuttum_button = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"p[class='text-decoration-none text-muted mt-5 d-block']"))).click()
      
        link_alani = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"input[class='form-control mt-6']")))
        link_alani.send_keys("s.kaya@msn.com")
        
        
        #link_alani = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='__next']/div/main/section/div/div/div/input"))).send_keys("s.kaya@msn.com")
        gonder_butonu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"button[class='btn btn-primary w-100 mt-6']"))).click()
        
        mesaj = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='toast-body']"))).text
        assert mesaj == "• Şifre sıfırlama linkini e-posta adresinize gönderdik. Lütfen gelen kutunuzu kontrol edin."
    
    def test_basarisiz_sifre_yenileme(self):
        sifremi_unuttum_button = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"p[class='text-decoration-none text-muted mt-5 d-block']"))).click()
      
        link_alani = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"input[class='form-control mt-6']"))).send_keys("@msn.com")
        #link_alani = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='__next']/div/main/section/div/div/div/input"))).send_keys("s.kaya@msn.com")
        gonder_butonu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"button[class='btn btn-primary w-100 mt-6']"))).click()
        mesaj =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='toast-body']"))).text
        assert mesaj == "• Girdiğiniz e-posta geçersizdir."
           

    def test_chatbot_mesaj_bolumu_acilma(self):
        
        iframe = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"iframe[class='exw-launcher-frame animated swing']")))
        WebDriverWait(self.driver,5).until(ec.frame_to_be_available_and_switch_to_it(iframe))
        
        chatbot_butonu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"button[id='launcher']")))
        chatbot_butonu.click()
        sleep(3)
        
        self.driver.switch_to.default_content()
        
        iframe_mesaj_kutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"iframe[class='exw-conversation-container-frame']")))
        WebDriverWait(self.driver,5).until(ec.frame_to_be_available_and_switch_to_it(iframe_mesaj_kutusu))
        
        tobeto_mesaj_kutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"h4[class='exw-title exw-with-avatar']")))
        var_mi = tobeto_mesaj_kutusu.is_displayed()
        assert var_mi == True
        
    
    def test_chatbot_kapatma(self):
    
        
        iframe = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"iframe[class='exw-launcher-frame animated swing']")))
        sleep(3)
        self.driver.switch_to.frame(iframe)
        
        
        chatbot_butonu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"button[id='launcher']")))
        chatbot_butonu.click()
        sleep(3)
        
        self.driver.switch_to.default_content()
        
        iframe_mesaj_kutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"iframe[class='exw-conversation-container-frame']")))
        WebDriverWait(self.driver,5).until(ec.frame_to_be_available_and_switch_to_it(iframe_mesaj_kutusu))
        simge_durumuna_getir_ikonu = self.driver.find_element(By.CSS_SELECTOR,"svg[class='exw-minimize-button header-button']").click()
        #simge= self.driver.find_element(By.XPATH,"//*[name()='svg' and @class='exw-minimize-button header-button']").click()
        sleep(3)

        self.driver.switch_to.default_content()
        iframe = WebDriverWait(self.driver,30).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"iframe[class='exw-launcher-frame animated swing']")))
        sleep(3)
        self.driver.switch_to.frame(iframe)
        
        chatbot_butonu = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"button[id='launcher']")))
    
        var_mi = chatbot_butonu.is_displayed()
        
        assert var_mi == True
    def test_kayitli_email_ile_kayitol(self):
        self.driver.execute_script("window.scrollBy(0,300)","")

        kayit_ol_buton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"a[class='text-decoration-none text-muted fw-bold']"))).click()
        
        sleep(3)
        self.driver.execute_script("window.scrollBy(0,200)","")
        ad = self.driver.find_element(By.NAME,"firstName").send_keys("ali")
        soyad = self.driver.find_element(By.NAME,"lastName").send_keys("kaya")
        email = self.driver.find_element(By.NAME,"email").send_keys("s.kaya@msn.com")
        sifre = self.driver.find_element(By.NAME,"password").send_keys("123456")
        sifretekrar = self.driver.find_element(By.NAME,"passwordAgain").send_keys("123456")
        self.driver.execute_script("window.scrollBy(0,200)","")
        sleep(2)
        kayitolbutton2 = self.driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-primary w-100 mt-6']").click()
        sleep(2)
        acıkrizametni = self.driver.find_element(By.NAME,"contact").click()
        uyeliksozlesmesi = self.driver.find_element(By.NAME,"membershipContrat").click()
        emailgonderimizni = self.driver.find_element(By.NAME,"emailConfirmation").click()
        aramaizni =self.driver.find_element(By.NAME,"phoneConfirmation").click()
        telno = self.driver.find_element(By.ID,"phoneNumber").send_keys("5555555555")
        iframe = self.driver.find_element(By.CSS_SELECTOR,"iframe[title='reCAPTCHA']")
        self.driver.switch_to.frame(iframe)
        sleep(2)
        robotdegilim = self.driver.find_element(By.ID,"recaptcha-anchor").click()
        sleep(10)
        self.driver.switch_to.default_content()
        devamet_butonu =self.driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-yes my-3']").click()
        sleep(3)
        mesaj = self.driver.find_element(By.CSS_SELECTOR,"div[class='toast-body']")
        assert mesaj.text == "• Girdiğiniz e-posta adresi ile kayıtlı üyelik bulunmaktadır."
    # chatbot case3 ve case4      
    def test_chatbot_uyari_mesaji_ve_puanlama_goruntulenme_kontrolu(self):
        self.test_chatbot_mesaj_bolumu_acilma()
        print("burda kalıyorum")
        sleep(3)
        # self.driver.switch_to.default_content()
        # iframe_mesajlasma_bitirme_ikonu = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"iframe[class='exw-conversation-container-frame']")))
        # WebDriverWait(self.driver,20).until(ec.frame_to_be_available_and_switch_to_it(iframe_mesajlasma_bitirme_ikonu))
        mesajlasma_bitirme_ikonu = self.driver.find_element(By.CSS_SELECTOR,"svg[class='exw-end-session-button header-button']")
        mesajlasma_bitirme_ikonu.click()
        sleep(3)
        sleep(5)
        bitirme_mesaji=WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='exw-finishSessionCheck']")))
        assert "Görüşmeyi bitirmek istediğinize emin misiniz?"  in bitirme_mesaji.text
        evet_butonu = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,"//button[text()='Evet']")))
        evet_butonu.click()
        puanlama = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='exw-survey-experience']")))
        assert puanlama.is_displayed() == True

